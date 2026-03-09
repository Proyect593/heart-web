from __future__ import annotations

import datetime as dt
import hashlib
import io
from pathlib import Path
from typing import Any

import firebase_admin
import openpyxl
from firebase_admin import credentials, db
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


class FirebaseService:
    """Encapsula la logica de acceso a Firebase y transformaciones de datos."""

    DB_URL = "https://heart-trazabilidad-default-rtdb.firebaseio.com/"

    def __init__(self, project_dir: Path) -> None:
        self.project_dir = Path(project_dir)
        self.root_dir = self.project_dir.parent
        self.cred_path = self.root_dir / "clave_firebase.json"

    def get_logo_path(self) -> Path:
        return self.root_dir / "logo.png"

    def _ensure_db(self):
        if not firebase_admin._apps:
            if not self.cred_path.exists():
                raise FileNotFoundError(f"No existe el archivo de credenciales: {self.cred_path}")
            cred = credentials.Certificate(str(self.cred_path))
            firebase_admin.initialize_app(cred, {"databaseURL": self.DB_URL})
        return db

    @staticmethod
    def _ordered_items(data_dict: dict[str, Any]) -> list[tuple[str, Any]]:
        def _sort_key(item: tuple[str, Any]) -> tuple[int, Any]:
            key = str(item[0])
            return (0, int(key)) if key.isdigit() else (1, key)

        return sorted(data_dict.items(), key=_sort_key)

    def _normalize_records(self, raw_data: Any) -> list[dict[str, str]]:
        items: list[tuple[str, Any]] = []
        if isinstance(raw_data, dict):
            if any(isinstance(v, dict) for v in raw_data.values()):
                items = self._ordered_items(raw_data)
            else:
                items = [("0", raw_data)]
        elif isinstance(raw_data, list):
            items = [(str(idx), value) for idx, value in enumerate(raw_data) if value]

        normalized: list[dict[str, str]] = []
        for key, record in items:
            if not isinstance(record, dict):
                continue
            normalized.append(
                {
                    "key": str(key),
                    "fecha": str(record.get("fecha", "")),
                    "hora": str(record.get("hora", "")),
                    "proceso": str(record.get("proceso", "")),
                    "punto": str(record.get("punto", "")),
                    "usuario": str(record.get("usuario", "")),
                }
            )
        return normalized

    def get_order_records(self, order_id: str) -> list[dict[str, str]]:
        db_mod = self._ensure_db()
        raw_data = db_mod.reference(f"/documentos/{order_id}").get()
        if not raw_data:
            return []
        return self._normalize_records(raw_data)

    def save_order_records(self, previous_order_id: str, new_order_id: str, records: list[dict[str, Any]]) -> int:
        if not new_order_id.strip():
            raise ValueError("El campo Documento no puede estar vacio.")
        if not records:
            raise ValueError("No hay registros para guardar.")

        payload: dict[str, dict[str, str]] = {}
        for index, record in enumerate(records):
            key = str(record.get("key", "")).strip() or str(index)
            payload[key] = {
                "fecha": str(record.get("fecha", "")).strip(),
                "hora": str(record.get("hora", "")).strip(),
                "proceso": str(record.get("proceso", "")).strip(),
                "punto": str(record.get("punto", "")).strip(),
                "usuario": str(record.get("usuario", "")).strip(),
            }

        db_mod = self._ensure_db()
        db_mod.reference(f"/documentos/{new_order_id}").set(payload)
        if new_order_id != previous_order_id:
            db_mod.reference(f"/documentos/{previous_order_id}").delete()

        return len(payload)

    def get_month_index(self) -> dict[str, set[str]]:
        db_mod = self._ensure_db()
        documentos = db_mod.reference("/documentos").get() or {}
        month_index: dict[str, set[str]] = {}

        if not isinstance(documentos, dict):
            return month_index

        for doc_id, registros in documentos.items():
            registros_iter: list[dict[str, Any]] = []
            if isinstance(registros, dict):
                if any(isinstance(v, dict) for v in registros.values()):
                    registros_iter = [v for v in registros.values() if isinstance(v, dict)]
                else:
                    registros_iter = [registros]
            elif isinstance(registros, list):
                registros_iter = [r for r in registros if isinstance(r, dict)]

            for reg in registros_iter:
                fecha = str(reg.get("fecha", "")).strip()
                if len(fecha) >= 7:
                    month = fecha[:7]
                    month_index.setdefault(month, set()).add(str(doc_id))

        return month_index

    def delete_months(self, months: list[str]) -> dict[str, int]:
        month_index = self.get_month_index()
        docs_to_delete: set[str] = set()
        for month in months:
            docs_to_delete.update(month_index.get(str(month), set()))

        db_mod = self._ensure_db()
        deleted = 0
        errors = 0
        for doc_id in docs_to_delete:
            try:
                db_mod.reference(f"/documentos/{doc_id}").delete()
                deleted += 1
            except Exception:
                errors += 1

        return {"selected": len(docs_to_delete), "deleted": deleted, "errors": errors}

    def delete_solucionadas(self) -> None:
        db_mod = self._ensure_db()
        db_mod.reference("/solucionadas").delete()

    def _get_firmas_ref_and_data(self):
        db_mod = self._ensure_db()
        ref = db_mod.reference("/Firmas_BPM")
        firmas = ref.get()
        if firmas is None:
            legacy_ref = db_mod.reference("/Clave/Firmas_BPM")
            legacy = legacy_ref.get()
            if legacy is not None:
                firmas = legacy
                ref.set(legacy)
            else:
                firmas = {}
        return ref, firmas

    @staticmethod
    def _find_user_by_document(firmas: Any, document: str) -> tuple[str | None, dict[str, Any] | None]:
        target = str(document).strip().lstrip("0")
        if isinstance(firmas, list):
            for idx, data in enumerate(firmas):
                candidate = str((data or {}).get("documento", "")).strip().lstrip("0")
                if candidate == target:
                    return str(idx), (data or {})
        elif isinstance(firmas, dict):
            for key, data in firmas.items():
                candidate = str((data or {}).get("documento", "")).strip().lstrip("0")
                if candidate == target:
                    return str(key), (data or {})
        return None, None

    @staticmethod
    def _find_user_by_signature(firmas: Any, signature: str) -> tuple[str | None, dict[str, Any] | None]:
        target = str(signature).strip().upper()
        if isinstance(firmas, list):
            for idx, data in enumerate(firmas):
                candidate = str((data or {}).get("firma", "")).strip().upper()
                if candidate == target:
                    return str(idx), (data or {})
        elif isinstance(firmas, dict):
            for key, data in firmas.items():
                candidate = str((data or {}).get("firma", "")).strip().upper()
                if candidate == target:
                    return str(key), (data or {})
        return None, None

    @staticmethod
    def _hash_password(password: str) -> str:
        return hashlib.sha256(password.encode()).hexdigest()

    def get_admin_credentials(self) -> dict[str, str] | None:
        db_mod = self._ensure_db()
        admin_data = db_mod.reference("/Admin").get()
        if not isinstance(admin_data, dict):
            return None

        username = str(admin_data.get("usuario") or admin_data.get("username") or admin_data.get("email") or "").strip()
        password = str(admin_data.get("contrasena") or admin_data.get("password") or "").strip()
        if not username or not password:
            return None

        return {"username": username, "password": password}

    def ensure_admin_credentials(self, username: str, password: str, overwrite: bool = False) -> dict[str, str]:
        clean_username = str(username).strip()
        clean_password = str(password).strip()

        if not clean_username or not clean_password:
            raise ValueError("El usuario y la contrasena de Admin son obligatorios.")

        db_mod = self._ensure_db()
        ref = db_mod.reference("/Admin")
        existing = ref.get()
        if isinstance(existing, dict) and not overwrite:
            existing_username = str(existing.get("usuario") or existing.get("username") or existing.get("email") or "").strip()
            existing_password = str(existing.get("contrasena") or existing.get("password") or "").strip()
            if existing_username and existing_password:
                return {"username": existing_username, "password": existing_password}

        payload = {
            "usuario": clean_username,
            "contrasena": clean_password,
        }
        ref.set(payload)
        return {"username": clean_username, "password": clean_password}

    def get_user_by_document(self, document: str) -> dict[str, str] | None:
        _, firmas = self._get_firmas_ref_and_data()
        user_key, user_data = self._find_user_by_document(firmas, document)
        if user_key is None or user_data is None:
            return None
        return {
            "key": user_key,
            "nombre": str(user_data.get("nombre", "")),
            "documento": str(user_data.get("documento", document)),
            "firma": str(user_data.get("firma", "")),
        }

    def update_user_password(self, document: str, password: str) -> None:
        if not password.isdigit() or len(password) != 4:
            raise ValueError("La nueva contrasena debe tener 4 digitos.")

        ref, firmas = self._get_firmas_ref_and_data()
        user_key, _ = self._find_user_by_document(firmas, document)
        if user_key is None:
            raise ValueError("No se encontro el usuario.")

        ref.child(str(user_key)).update({"contrasena": self._hash_password(password), "contrase\u00f1a": None})

    def update_clave_general(self, new_value: str) -> None:
        if not str(new_value).strip():
            raise ValueError("La clave nueva no puede estar vacia.")
        db_mod = self._ensure_db()
        db_mod.reference("/Clave").child("1").set(str(new_value).strip())

    def reset_user_password(self, document: str) -> dict[str, str]:
        ref, firmas = self._get_firmas_ref_and_data()
        user_key, user_data = self._find_user_by_document(firmas, document)
        if user_key is None or user_data is None:
            raise ValueError("No se encontro el usuario.")

        ref.child(str(user_key)).update({"contrasena": "", "contrase\u00f1a": None})
        return {
            "key": str(user_key),
            "nombre": str(user_data.get("nombre", "")),
            "documento": str(user_data.get("documento", document)),
        }

    def create_user(self, nombre: str, documento: str, firma: str) -> dict[str, str | int]:
        nombre = str(nombre).strip().upper()
        documento = str(documento).strip()
        firma = str(firma).strip().upper()

        if not nombre or not documento or not firma:
            raise ValueError("Completa todos los campos.")

        ref, firmas = self._get_firmas_ref_and_data()

        existing_doc_key, _ = self._find_user_by_document(firmas, documento)
        if existing_doc_key is not None:
            raise ValueError("Ya existe un usuario con ese documento.")

        existing_sig_key, _ = self._find_user_by_signature(firmas, firma)
        if existing_sig_key is not None:
            raise ValueError("Ya existe un usuario con esa firma BPM.")

        if isinstance(firmas, dict) and firmas:
            numeric_keys = [int(str(k)) for k in firmas.keys() if str(k).isdigit()]
            next_index = (max(numeric_keys) + 1) if numeric_keys else len(firmas)
            next_key = str(next_index)
            record_number = next_index + 1
        elif isinstance(firmas, list):
            next_index = len(firmas)
            next_key = str(next_index)
            record_number = next_index + 1
        else:
            next_key = "0"
            record_number = 1

        new_user = {
            "nombre": nombre,
            "documento": documento,
            "firma": firma,
            "numero": record_number,
        }

        ref.child(next_key).set(new_user)
        return {"key": next_key, **new_user}

    @staticmethod
    def _extract_comments(comment_block: Any) -> list[dict[str, str]]:
        extracted: list[dict[str, str]] = []

        def _from_comment(raw_comment: Any) -> None:
            if isinstance(raw_comment, dict):
                comment_text = str(raw_comment.get("comentario", "")).strip()
                fecha_hora = str(raw_comment.get("fecha", "")).strip()
                fecha_comment = ""
                hora_comment = ""
                if fecha_hora:
                    if " " in fecha_hora:
                        fecha_comment, hora_comment = fecha_hora.split(" ", 1)
                    else:
                        fecha_comment = fecha_hora
                if comment_text:
                    extracted.append(
                        {
                            "comentario": comment_text,
                            "fecha_comentario": fecha_comment,
                            "hora_comentario": hora_comment,
                        }
                    )
            elif isinstance(raw_comment, str):
                comment_text = raw_comment.strip()
                if comment_text:
                    extracted.append(
                        {
                            "comentario": comment_text,
                            "fecha_comentario": "",
                            "hora_comentario": "",
                        }
                    )

        if isinstance(comment_block, dict):
            for _, item in comment_block.items():
                _from_comment(item)
        elif isinstance(comment_block, list):
            for item in comment_block:
                _from_comment(item)

        return extracted

    def build_backup_excel(self) -> tuple[bytes, str]:
        db_mod = self._ensure_db()
        data = db_mod.reference("/solucionadas").get()
        if not data:
            raise ValueError("No se encontraron novedades solucionadas.")

        columns = [
            "Consecutivo",
            "Descripcion",
            "Punto",
            "Fecha",
            "Comentario",
            "Fecha comentario",
            "Hora comentario",
            "Solucionada",
        ]

        rows: list[dict[str, str]] = []
        for _, node in data.items():
            consecutivo = str(node.get("consecutivo", ""))
            descripcion = str(node.get("descripcion", ""))
            punto = str(node.get("punto", ""))
            fecha = str(node.get("fecha", ""))
            solucionada = bool(node.get("solucionada", False))
            comentarios = node.get("comentarios")

            for comment in self._extract_comments(comentarios):
                rows.append(
                    {
                        "Consecutivo": consecutivo,
                        "Descripcion": descripcion,
                        "Punto": punto,
                        "Fecha": fecha,
                        "Comentario": comment["comentario"],
                        "Fecha comentario": comment["fecha_comentario"],
                        "Hora comentario": comment["hora_comentario"],
                        "Solucionada": "Si" if solucionada else "No",
                    }
                )

        rows.sort(key=lambda item: item["Fecha"], reverse=True)

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Historial novedades"
        sheet.append(columns)

        for row in rows:
            sheet.append([row[col] for col in columns])

        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        thin = Side(border_style="thin", color="000000")
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col in sheet.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            sheet.column_dimensions[col_letter].width = max_length + 4

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        timestamp = dt.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        filename = f"Historial_novedades_{timestamp}.xlsx"
        return output.getvalue(), filename
